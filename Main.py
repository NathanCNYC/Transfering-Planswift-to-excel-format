import os
import shutil
import pandas as pd
from openpyxl import load_workbook
import re

from refine import refine_values  # Import refine_values function
from populate_calculator import populate_piping, populate_ductwork  # Import populate functions


def is_date(value):
    """
    Checks if a value is a date in the format MM/DD/YYYY.
    Args:
        value (str): The value to check.
    Returns:
        bool: True if the value matches the date format, otherwise False.
    """
    date_pattern = r'^\d{1,2}/\d{1,2}/\d{4}$'  # Matches MM/DD/YYYY
    if isinstance(value, str) and re.match(date_pattern, value):
        return True
    return False


def process_csv_to_excel(data_file, template_file, output_file):
    """
    Processes a CSV file to copy rows into the 'raw input' sheet of a template file.
    Fills empty cells with placeholders (e.g., 0 for numeric fields, '-' for text fields).

    Args:
        data_file (str): Path to the CSV file containing the data.
        template_file (str): Path to the Excel file template (.xlsx).
        output_file (str): Path to save the modified output file (.xlsx).
    """
    try:
        # Copy the template file to the output file
        print(f"Copying template file to {output_file}...")
        shutil.copy(template_file, output_file)

        # Read the CSV file, handle rows with varying columns
        print("Reading the CSV file...")
        data = pd.read_csv(data_file, header=None, on_bad_lines="skip", engine="python")

        # Ensure all rows have exactly 9 columns
        print("Trimming and padding rows with placeholders...")
        data = data.apply(
            lambda x: [val if pd.notna(val) and val != "" else (0 if i in [1, 3, 6, 7, 8] else "-") for i, val in enumerate(x[:9])]
            + [""] * (9 - len(x)),
            axis=1,
        )

        # Convert the adjusted data back to a DataFrame
        data = pd.DataFrame(data.tolist())

        # Open the output Excel file
        print("Loading the output file...")
        workbook_out = load_workbook(output_file)

        # Create a new worksheet named 'raw input' as the first sheet
        if 'raw input' in workbook_out.sheetnames:
            print("Worksheet 'raw input' already exists. Overwriting...")
            workbook_out.remove(workbook_out['raw input'])
        sheet_out = workbook_out.create_sheet(title='raw input')

        # Write the header in the output file
        header = ['Name', 'Qty', 'Units', 'Square feet', 'Floor', 'Manufacturer', 'Labor', 'Equipment Total', 'Counts']
        print("Adding header...")
        sheet_out.append(header)

        # Process rows in the cleaned DataFrame
        print("Processing rows...")
        for row in data.itertuples(index=False):
            sheet_out.append(row)

        # Remove entirely empty rows based on Column A
        print("Removing empty rows from the output...")
        row_num = 2  # Start after the header row
        while row_num <= sheet_out.max_row:
            if not sheet_out.cell(row=row_num, column=1).value:  # Check Column A
                sheet_out.delete_rows(row_num)
            else:
                row_num += 1

        # Save the modified output file
        print(f"Saving changes to {output_file}...")
        workbook_out.save(output_file)
        print("Process completed successfully!")

    except Exception as e:
        print(f"An error occurred: {e}")




# Example usage
if __name__ == "__main__":
    data_file = "input.csv"  # Path to the raw data CSV file
    template_file = r"C:...." #<---- INPUT DIRECTORY OF YOUR TEMPLATE HERE
    output_file = "M Breakdown.xlsx"  # Path to the local output file

    # Step 1: Process input data
    process_csv_to_excel(data_file, template_file, output_file)

    # Step 2: Refine values
    refine_values(output_file)

    # Step 3: Populate calculator
    populate_piping(output_file)
    populate_ductwork(output_file)
