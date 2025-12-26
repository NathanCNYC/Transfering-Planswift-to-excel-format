import pandas as pd
from openpyxl import load_workbook

def refine_values(output_file):
    """
    Processes the 'raw input' sheet of an Excel workbook to identify duplicate entries,
    group them by similarity, and sum their Qty and Square Feet values.
    The refined data is saved in a new sheet called 'Refined values'.

    Args:
        output_file (str): Path to the Excel file where the 'raw input' sheet exists and
                          where the refined values will be written.
    """
    try:
        # Load the workbook and the 'raw input' sheet
        print("Loading the workbook...")
        workbook = load_workbook(output_file)
        if 'raw input' not in workbook.sheetnames:
            raise ValueError("'raw input' sheet not found in the workbook.")
        
        # Read the 'raw input' sheet into a DataFrame
        print("Reading 'raw input' sheet...")
        raw_input_df = pd.DataFrame(workbook['raw input'].values)
        
        # Use the first row as the header and drop the empty rows
        raw_input_df.columns = raw_input_df.iloc[0]
        raw_input_df = raw_input_df[1:].dropna(how='all')

        # Ensure the required columns exist
        required_columns = ['Name', 'Qty', 'Units', 'Square feet']
        for col in required_columns:
            if col not in raw_input_df.columns:
                raise ValueError(f"Column '{col}' not found in the 'raw input' sheet.")

        # Automatically clean and convert numbers stored as text to numeric
        print("Cleaning and converting numeric columns...")
        raw_input_df['Qty'] = pd.to_numeric(raw_input_df['Qty'].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        raw_input_df['Square feet'] = pd.to_numeric(raw_input_df['Square feet'].astype(str).str.replace(',', ''), errors='coerce').fillna(0)

        # Group by 'Name' and 'Units', and sum the values of 'Qty' and 'Square feet'
        print("Refining and grouping data...")
        refined_df = raw_input_df.groupby(['Name', 'Units'], as_index=False).agg({
            'Qty': 'sum',
            'Square feet': 'sum'
        })

        # Add a new sheet for the refined values
        print("Creating 'Refined values' sheet...")
        if 'Refined values' in workbook.sheetnames:
            del workbook['Refined values']
        refined_sheet = workbook.create_sheet('Refined values', index=1)

        # Write the header to the new sheet
        header = ['Name', 'Units', 'Total Qty', 'Total Square feet']
        refined_sheet.append(header)

        # Write the refined data to the new sheet
        print("Writing refined data...")
        for _, row in refined_df.iterrows():
            refined_sheet.append([row['Name'], row['Units'], row['Qty'], row['Square feet']])

        # Save the workbook
        print(f"Saving changes to {output_file}...")
        workbook.save(output_file)
        print("Process completed successfully!")

    except Exception as e:
        print(f"An error occurred: {e}")

# Example usage
if __name__ == "__main__":
    output_excel_path = "output.xlsx"  # Path to the Excel file

    refine_values(output_excel_path)
