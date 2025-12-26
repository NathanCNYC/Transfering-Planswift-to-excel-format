import openpyxl
import re

def populate_piping(output_file):
    """
    Populates predefined areas in the 'Mechanical Breakdown' sheet for piping sections (Refrigerant and Condensate Drain)
    using data from the 'Refined values' sheet in the same workbook. It looks for the 'INPUT AREA' marker and restricts
    size detection to that column.

    Args:
        output_file (str): Path to the output Excel file containing both 'Refined values' and 'Mechanical Breakdown' sheets.
    """
    try:
        print("Loading workbook...")
        workbook = openpyxl.load_workbook(output_file)

        if 'Refined values' not in workbook.sheetnames:
            raise ValueError("'Refined values' sheet not found in the workbook.")

        if 'Mechanical Breakdown' not in workbook.sheetnames:
            raise ValueError("'Mechanical Breakdown' sheet not found in the workbook.")

        refined_sheet = workbook['Refined values']
        breakdown_sheet = workbook['Mechanical Breakdown']

        # Locate the 'INPUT AREA' cell to restrict size detection to its column
        print("Locating 'INPUT AREA'...")
        input_area_column = None
        for row in breakdown_sheet.iter_rows(min_row=1, max_col=breakdown_sheet.max_column, values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().upper() == "INPUT AREA":
                    input_area_column = cell.column
                    print(f"'INPUT AREA' found in column {input_area_column}.")
                    break
            if input_area_column:
                break

        if input_area_column is None:
            raise ValueError("'INPUT AREA' not found in the 'Mechanical Breakdown' sheet.")

        # Process Refrigerant Piping
        print("Processing refrigerant piping...")
        refrigerant_value = None
        for row in refined_sheet.iter_rows(min_row=2, values_only=True):
            name, _, qty, *_ = row  # Skip the second column
            if isinstance(name, str) and "refrigerant" in name.lower():
                refrigerant_value = qty
                break

        if refrigerant_value is not None:
            for row in breakdown_sheet.iter_rows(min_row=1, max_col=breakdown_sheet.max_column):
                for cell in row:
                    if isinstance(cell.value, str) and "Input Refrigerant Piping" in cell.value:
                        target_cell = breakdown_sheet.cell(row=cell.row, column=cell.column + 1)
                        target_cell.value = refrigerant_value
                        print(f"Refrigerant piping value '{refrigerant_value}' placed in cell {target_cell.coordinate}.")
                        break

        # Process Condensate Drain
        print("Processing condensate drain...")
        size_column = {
            str(breakdown_sheet.cell(row=row, column=input_area_column).value).strip(): row
            for row in range(1, breakdown_sheet.max_row + 1)
            if isinstance(breakdown_sheet.cell(row=row, column=input_area_column).value, str)
        }

        for row in refined_sheet.iter_rows(min_row=2, values_only=True):
            name, _, qty, *_ = row  # Skip the second column
            if isinstance(name, str) and "condensate drain" in name.lower():
                # Match the full size string, including fractions and quotes
                match = re.search(r'(\d+-\d+/\d+|\d+/\d+|\d+-\d+|\d+)', name)
                if match:
                    size = match.group(0).strip()
                    size = size + '"' if '"' not in size else size  # Ensure size includes the quote
                    if size in size_column:
                        target_row = size_column[size]
                        target_cell = breakdown_sheet.cell(row=target_row, column=input_area_column + 1)  # Place value in the next column
                        target_cell.value = qty
                        print(f"Condensate drain value '{qty}' placed in row {target_row}, column {input_area_column + 1} for size '{size}'.")
                    else:
                        print(f"Size '{size}' not found in column {input_area_column}.")
                else:
                    print(f"No valid size found in: {name}")
        
        # Process Copper
        print("Processing Copper...")
        for row in refined_sheet.iter_rows(min_row=2, values_only=True):
            name, _, qty, *_ = row
            if isinstance(name, str) and "equipment riser & branch piping" in name.lower():
                match = re.search(r'(\d+-\d+/\d+|\d+/\d+|\d+-\d+|\d+)', name)
                if match:
                    size = match.group(0).strip()
                    size = size + '"' if '"' not in size else size  # Ensure size includes the quote
                    if size in size_column:
                        target_row = size_column[size]
                        target_cell = breakdown_sheet.cell(row=target_row, column=input_area_column + 2)  # Copper column
                        target_cell.value = qty
                        print(f"Copper value '{qty}' placed in row {target_row}, column {input_area_column + 2} for size '{size}'.")
                    else:
                        print(f"Size '{size}' not found in column {input_area_column} for Copper.")
                else:
                    print(f"No valid size found in: {name}")

        # Process SCH40
        print("Processing SCH40...")
        for row in refined_sheet.iter_rows(min_row=2, values_only=True):
            name, _, qty, *_ = row
            if isinstance(name, str) and "sch 40 blk iron" in name.lower():
                match = re.search(r'(\d+-\d+/\d+|\d+/\d+|\d+-\d+|\d+)', name)
                if match:
                    size = match.group(0).strip()
                    size = size + '"' if '"' not in size else size  # Ensure size includes the quote
                    if size in size_column:
                        target_row = size_column[size]
                        target_cell = breakdown_sheet.cell(row=target_row, column=input_area_column + 3)  # SCH40 column
                        target_cell.value = qty
                        print(f"SCH40 value '{qty}' placed in row {target_row}, column {input_area_column + 3} for size '{size}'.")
                    else:
                        print(f"Size '{size}' not found in column {input_area_column} for SCH40.")
                else:
                    print(f"No valid size found in: {name}")


        # Save the updated workbook
        print(f"Saving changes to {output_file}...")
        workbook.save(output_file)
        print("Piping section populated successfully!")

    except Exception as e:
        print(f"An error occurred: {e}")

def populate_ductwork(output_file):
    """
    Populates the 'Mechanical Breakdown' sheet for the ductwork section by summing all values
    for specified categories in the 'Refined values' sheet and placing the totals in predefined cells.

    This version handles both:
      - Standard categories (Galv, Kitchen, Aluminum, Flat Oval) in the "Qty" column.
      - Acoustically Lined / Insulated / Fire-Wrapped combos for each base category (Galv, Kitchen, etc.) 
        in the "Square feet" column, feeding into the appropriate mechanical breakdown cells.

    If a cell already contains a number, new sums are added (accumulated).
    """
    try:
        print("Loading workbook...")
        workbook = openpyxl.load_workbook(output_file)

        if 'Refined values' not in workbook.sheetnames:
            raise ValueError("'Refined values' sheet not found in the workbook.")

        if 'Mechanical Breakdown' not in workbook.sheetnames:
            raise ValueError("'Mechanical Breakdown' sheet not found in the workbook.")

        refined_sheet = workbook['Refined values']
        breakdown_sheet = workbook['Mechanical Breakdown']

        # ----------------------------------------------------------------------
        # 1) Summarize standard ductwork categories in the 3rd column (Qty)
        # ----------------------------------------------------------------------
        print("Summing category values for standard ductwork types...")
        base_categories = {
            "Galvanized Steel": "Input Galvanized Steel",
            "Residential Kitchen": "Input Residential Kitchen",
            "Commercial Kitchen": "Input Commercial Kitchen",
            "Aluminum": "Input Aluminum",
            "Flat Oval": " Input Flat Oval",
            "316 SS 18 Gauge DX": " (Usually Ignore) Stainless Steel"

        }

        # aggregator for base categories (from the "Qty" column)
        base_aggregator = {cat: 0 for cat in base_categories}

        # Go through "Refined values" to sum base categories
        for row in refined_sheet.iter_rows(min_row=2, values_only=True):
            name, _, qty, *_ = row  # skip second column, etc.
            if isinstance(name, str) and isinstance(qty, (int, float)):
                for cat in base_categories:
                    # If the category name is present in 'name'
                    if cat.lower() in name.lower():
                        base_aggregator[cat] += qty

        print("Placing ductwork type totals in Mechanical Breakdown (accumulating if needed)...")
        for cat, label in base_categories.items():
            new_value = base_aggregator[cat]
            # Search for cells with the matching label
            for row in breakdown_sheet.iter_rows(min_row=1, max_col=breakdown_sheet.max_column):
                for cell in row:
                    if isinstance(cell.value, str) and label in cell.value:
                        # The cell to the right is where the numeric total goes
                        target_cell = breakdown_sheet.cell(row=cell.row, column=cell.column + 1)
                        # Accumulate if there's already a value
                        existing_val = target_cell.value
                        if not isinstance(existing_val, (int, float)):
                            existing_val = 0
                        target_cell.value = existing_val + new_value
                        print(f"{cat} total '{new_value}' added to cell {target_cell.coordinate}. "
                              f"(Previous: {existing_val}, New: {existing_val + new_value})")
                        break

        # ----------------------------------------------------------------------
        # 2) Summarize acoustically lined / insulated / fire-wrapped combos for each base category
        #    in the 4th column (Square Feet)
        # ----------------------------------------------------------------------
        print("Summarizing Acoustically Lined / Insulated / Fire-Wrapped combos...")

        # Build a dictionary of combos => cell label in Mechanical Breakdown
        # e.g., "Acoustically Lined Galvanized Steel" => "Input Acoustical Lining (SqFt)"
        # define "Acoustically Lined *Category*" => same label,
        # "Insulated *Category*" => same label, etc.
        sq_categories = {
            # For each base, define the combos
            "Acoustically Lined Galvanized Steel": "Input Acoustical Lining (SqFt)",
            "Insulated Galvanized Steel": "Input Insulation (SqFt)",
            "Fire Wrapped Galvanized Steel": "Input Fire Wrapped (SqFt)",

            "Acoustically Lined Residential Kitchen": "Input Acoustical Lining (SqFt)",
            "Insulated Residential Kitchen": "Input Insulation (SqFt)",
            "Fire Wrapped Residential Kitchen": "Input Fire Wrapped (SqFt)",

            "Acoustically Lined Commercial Kitchen": "Input Acoustical Lining (SqFt)",
            "Insulated Commercial Kitchen": "Input Insulation (SqFt)",
            "Fire Wrapped Commercial Kitchen": "Input Fire Wrapped (SqFt)",

            "Acoustically Lined Aluminum": "Input Acoustical Lining (SqFt)",
            "Insulated Aluminum": "Input Insulation (SqFt)",
            "Fire Wrapped Aluminum": "Input Fire Wrapped (SqFt)",

            "Acoustically Lined Flat Oval": "Input Acoustical Lining (SqFt)",
            "Insulated Flat Oval": "Input Insulation (SqFt)",
            "Fire Wrapped Flat Oval": "Input Fire Wrapped (SqFt)",

            "Acoustically Lined 316 SS 18 Gauge DX": "Input Acoustical Lining (SqFt)",
            "Insulated 316 SS 18 Gauge DX": "Input Insulation (SqFt)",
            "Fire Wrapped 316 SS 18 Gauge DX": "Input Fire Wrapped (SqFt)",
        }

        # aggregator for combos from the "Square feet" column (4th col)
        sq_aggregator = {combo: 0 for combo in sq_categories}

        # Summation from "Refined values" (4th column => sq_ft)
        for row in refined_sheet.iter_rows(min_row=2, values_only=True):
            name, _, _, sq_ft, *_ = row
            if isinstance(name, str) and isinstance(sq_ft, (int, float)):
                # Check each combo
                for combo in sq_categories:
                    if combo.lower() in name.lower():
                        sq_aggregator[combo] += sq_ft

        print("Placing acoustical/insulation/fire-wrapped SQFt combos in Mechanical Breakdown (accumulating if needed)...")
        for combo, label in sq_categories.items():
            new_sq_value = sq_aggregator[combo]
            # Find label in mechanical breakdown
            for row in breakdown_sheet.iter_rows(min_row=1, max_col=breakdown_sheet.max_column):
                for cell in row:
                    if isinstance(cell.value, str) and label in cell.value:
                        # read existing cell
                        target_cell = breakdown_sheet.cell(row=cell.row, column=cell.column + 1)
                        existing_val = target_cell.value
                        if not isinstance(existing_val, (int, float)):
                            existing_val = 0
                        target_cell.value = existing_val + new_sq_value
                        print(f"{combo} => '{new_sq_value}' added to {label} cell {target_cell.coordinate}. "
                              f"(Previous: {existing_val}, New: {existing_val + new_sq_value})")
                        break

        # ----------------------------------------------------------------------
        # Save the updated workbook
        # ----------------------------------------------------------------------
        print(f"Saving changes to {output_file}...")
        workbook.save(output_file)
        print("Ductwork section populated successfully!")

    except Exception as e:
        print(f"An error occurred: {e}")


# Example usage
if __name__ == "__main__":
    output_file_path = "M Breakdown.xlsx"  # Path to the workbook containing both sheets
    populate_piping(output_file_path)
    populate_ductwork(output_file_path)
